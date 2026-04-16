import streamlit as st
import pandas as pd
import re
import io
import random
from faker import Faker
from openpyxl.styles import PatternFill, Font

# ─────────────────────────────────────────────
#  Setup
#  NOTE: Streamlit theme is now configured via a real
#  `.streamlit/config.toml` file in the repo — NOT written
#  at runtime. Writing config.toml at runtime has no effect
#  (only read on startup) and breaks read-only deployments.
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="DataMask | Professional Data Anonymization",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
#  PROFESSIONAL DARK THEME — Custom CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
/* ══════════════════════════════════════════════
   IMPORTS — Premium Typography
   ══════════════════════════════════════════════ */
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,400&family=JetBrains+Mono:wght@400;500&display=swap');

/* ══════════════════════════════════════════════
   CSS VARIABLES
   ══════════════════════════════════════════════ */
:root {
    --bg-primary: #0B0F19;
    --bg-secondary: #111827;
    --bg-card: #1A1F2E;
    --bg-card-hover: #212840;
    --bg-input: #151A28;
    --border-subtle: rgba(255,255,255,0.06);
    --border-accent: rgba(56, 189, 248, 0.2);
    --text-primary: #F1F5F9;
    --text-secondary: #94A3B8;
    /* FIX: bumped from #64748B (3.45:1) to #94A3B8 (7.47:1) to meet WCAG AA */
    --text-muted: #94A3B8;
    --accent-cyan: #38BDF8;
    --accent-emerald: #34D399;
    --accent-violet: #A78BFA;
    --accent-amber: #FBBF24;
    --accent-rose: #FB7185;
    --glow-cyan: rgba(56, 189, 248, 0.15);
    --glow-emerald: rgba(52, 211, 153, 0.15);
}

/* ══════════════════════════════════════════════
   GLOBAL RESET & BASE
   ══════════════════════════════════════════════ */
html, body, .stApp, [data-testid="stAppViewContainer"] {
    background-color: var(--bg-primary) !important;
    color: var(--text-primary) !important;
    font-family: 'DM Sans', sans-serif !important;
}

.stApp {
    background: var(--bg-primary) !important;
}

/* Animated gradient mesh background */
.stApp::before {
    content: '';
    position: fixed;
    top: 0; left: 0; right: 0; bottom: 0;
    background:
        radial-gradient(ellipse 80% 50% at 20% 40%, rgba(56, 189, 248, 0.06), transparent),
        radial-gradient(ellipse 60% 40% at 80% 20%, rgba(167, 139, 250, 0.05), transparent),
        radial-gradient(ellipse 50% 60% at 60% 80%, rgba(52, 211, 153, 0.04), transparent);
    pointer-events: none;
    z-index: 0;
    animation: meshShift 20s ease-in-out infinite alternate;
}

@keyframes meshShift {
    0% {
        background:
            radial-gradient(ellipse 80% 50% at 20% 40%, rgba(56, 189, 248, 0.06), transparent),
            radial-gradient(ellipse 60% 40% at 80% 20%, rgba(167, 139, 250, 0.05), transparent),
            radial-gradient(ellipse 50% 60% at 60% 80%, rgba(52, 211, 153, 0.04), transparent);
    }
    50% {
        background:
            radial-gradient(ellipse 70% 60% at 50% 30%, rgba(167, 139, 250, 0.07), transparent),
            radial-gradient(ellipse 80% 50% at 20% 70%, rgba(56, 189, 248, 0.05), transparent),
            radial-gradient(ellipse 60% 40% at 80% 50%, rgba(251, 191, 36, 0.04), transparent);
    }
    100% {
        background:
            radial-gradient(ellipse 60% 50% at 70% 60%, rgba(52, 211, 153, 0.06), transparent),
            radial-gradient(ellipse 80% 40% at 30% 20%, rgba(56, 189, 248, 0.05), transparent),
            radial-gradient(ellipse 50% 60% at 50% 40%, rgba(167, 139, 250, 0.05), transparent);
    }
}

/* ══════════════════════════════════════════════
   ANIMATED FLOATING PARTICLES (pure CSS)
   ══════════════════════════════════════════════ */
.particles-container {
    position: fixed;
    top: 0; left: 0; right: 0; bottom: 0;
    pointer-events: none;
    z-index: 0;
    overflow: hidden;
}

.particle {
    position: absolute;
    width: 3px;
    height: 3px;
    border-radius: 50%;
    opacity: 0;
    animation: floatUp linear infinite;
}

.particle:nth-child(1) { left: 10%; animation-duration: 18s; animation-delay: 0s; background: var(--accent-cyan); }
.particle:nth-child(2) { left: 25%; animation-duration: 22s; animation-delay: 3s; background: var(--accent-violet); }
.particle:nth-child(3) { left: 40%; animation-duration: 16s; animation-delay: 1s; background: var(--accent-emerald); }
.particle:nth-child(4) { left: 55%; animation-duration: 24s; animation-delay: 5s; background: var(--accent-cyan); }
.particle:nth-child(5) { left: 70%; animation-duration: 20s; animation-delay: 2s; background: var(--accent-violet); }
.particle:nth-child(6) { left: 85%; animation-duration: 19s; animation-delay: 4s; background: var(--accent-emerald); }
.particle:nth-child(7) { left: 15%; animation-duration: 21s; animation-delay: 6s; background: var(--accent-amber); width: 2px; height: 2px; }
.particle:nth-child(8) { left: 60%; animation-duration: 17s; animation-delay: 1.5s; background: var(--accent-rose); width: 2px; height: 2px; }
.particle:nth-child(9) { left: 35%; animation-duration: 23s; animation-delay: 7s; background: var(--accent-cyan); width: 4px; height: 4px; }
.particle:nth-child(10) { left: 78%; animation-duration: 25s; animation-delay: 3.5s; background: var(--accent-violet); width: 2px; height: 2px; }

@keyframes floatUp {
    0%   { transform: translateY(100vh) scale(0); opacity: 0; }
    10%  { opacity: 0.6; }
    90%  { opacity: 0.6; }
    100% { transform: translateY(-10vh) scale(1); opacity: 0; }
}

/* ══════════════════════════════════════════════
   ACCESSIBILITY: respect reduced-motion preference
   ══════════════════════════════════════════════ */
@media (prefers-reduced-motion: reduce) {
    .particle,
    .stApp::before,
    .feature-card,
    h1,
    .hero-subtitle,
    .trust-banner,
    .detect-table {
        animation: none !important;
        transition: none !important;
    }
    .feature-card:hover {
        transform: none !important;
    }
}

/* ══════════════════════════════════════════════
   MAIN CONTENT AREA
   ══════════════════════════════════════════════ */
.block-container {
    padding-top: 2.5rem !important;
    padding-bottom: 3rem !important;
    max-width: 1200px !important;
    position: relative;
    z-index: 1;
}

/* ══════════════════════════════════════════════
   TYPOGRAPHY
   ══════════════════════════════════════════════ */
h1 {
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 700 !important;
    letter-spacing: -1px !important;
    background: linear-gradient(135deg, #38BDF8 0%, #A78BFA 50%, #34D399 100%) !important;
    -webkit-background-clip: text !important;
    -webkit-text-fill-color: transparent !important;
    background-clip: text !important;
    font-size: 2.4rem !important;
    line-height: 1.2 !important;
    margin-bottom: 0.3rem !important;
    animation: fadeSlideIn 0.8s ease-out;
}

h2 {
    font-family: 'DM Sans', sans-serif !important;
    color: var(--text-primary) !important;
    font-weight: 600 !important;
    font-size: 1.3rem !important;
    letter-spacing: -0.3px !important;
}

h3 {
    font-family: 'DM Sans', sans-serif !important;
    color: var(--text-primary) !important;
    font-weight: 600 !important;
    font-size: 1.15rem !important;
}

p, li, label {
    font-family: 'DM Sans', sans-serif !important;
    color: var(--text-secondary) !important;
}

/* Apply font to markdown content only (avoids breaking widget icons) */
[data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] span,
[data-testid="stMarkdownContainer"] li,
[data-testid="stCaptionContainer"] p,
.stAlert p, .stAlert span {
    font-family: 'DM Sans', sans-serif !important;
}

/* ══════════════════════════════════════════════
   ENTRANCE ANIMATIONS
   ══════════════════════════════════════════════ */
@keyframes fadeSlideIn {
    from { opacity: 0; transform: translateY(16px); }
    to   { opacity: 1; transform: translateY(0); }
}

@keyframes fadeIn {
    from { opacity: 0; }
    to   { opacity: 1; }
}

@keyframes scaleIn {
    from { opacity: 0; transform: scale(0.95); }
    to   { opacity: 1; transform: scale(1); }
}

@keyframes slideRight {
    from { opacity: 0; transform: translateX(-20px); }
    to   { opacity: 1; transform: translateX(0); }
}

@keyframes glowPulse {
    0%, 100% { box-shadow: 0 0 20px rgba(56, 189, 248, 0.1); }
    50%      { box-shadow: 0 0 30px rgba(56, 189, 248, 0.2); }
}

/* ══════════════════════════════════════════════
   HERO HEADER AREA
   ══════════════════════════════════════════════ */
.hero-subtitle {
    font-size: 1.05rem;
    color: var(--text-secondary);
    line-height: 1.7;
    animation: fadeSlideIn 0.8s ease-out 0.15s both;
    margin-bottom: 1rem;
}

/* Trust Banner — glassmorphism */
.trust-banner {
    background: rgba(56, 189, 248, 0.06);
    border: 1px solid rgba(56, 189, 248, 0.15);
    backdrop-filter: blur(12px);
    -webkit-backdrop-filter: blur(12px);
    border-radius: 14px;
    padding: 1rem 1.8rem;
    text-align: center;
    margin: 1.2rem 0 1.8rem 0;
    animation: fadeSlideIn 0.8s ease-out 0.3s both;
    display: flex;
    justify-content: center;
    gap: 2rem;
    flex-wrap: wrap;
}

.trust-badge {
    display: inline-flex;
    align-items: center;
    gap: 0.4rem;
    font-size: 0.85rem;
    color: var(--text-secondary);
    font-weight: 500;
    letter-spacing: 0.2px;
}

.trust-badge .icon {
    font-size: 1rem;
}

/* ══════════════════════════════════════════════
   SIDEBAR
   ══════════════════════════════════════════════ */
[data-testid="stSidebar"] {
    background: var(--bg-secondary) !important;
    border-right: 1px solid var(--border-subtle) !important;
}

[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] span,
[data-testid="stSidebar"] label {
    color: var(--text-secondary) !important;
}

[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
    color: var(--accent-cyan) !important;
    font-size: 1rem !important;
    text-transform: uppercase;
    letter-spacing: 1.5px !important;
    font-weight: 600 !important;
}

/* Smaller group headings inside sidebar */
[data-testid="stSidebar"] .group-label {
    color: var(--text-muted) !important;
    font-size: 0.72rem !important;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    font-weight: 600;
    margin: 0.8rem 0 0.3rem 0 !important;
    display: block;
}

[data-testid="stSidebar"] .stDivider {
    border-color: var(--border-subtle) !important;
}

/* Sidebar checkbox styling */
[data-testid="stSidebar"] [data-testid="stCheckbox"] label span {
    color: var(--text-secondary) !important;
    font-size: 0.88rem !important;
}

[data-testid="stSidebar"] [data-testid="stCheckbox"] label:hover span {
    color: var(--text-primary) !important;
}

/* Radio buttons in sidebar */
[data-testid="stSidebar"] [role="radiogroup"] label {
    background: var(--bg-card) !important;
    border: 1px solid var(--border-subtle) !important;
    border-radius: 8px !important;
    padding: 0.5rem 0.8rem !important;
    transition: all 0.25s ease !important;
}

[data-testid="stSidebar"] [role="radiogroup"] label:hover {
    border-color: var(--accent-cyan) !important;
    background: var(--bg-card-hover) !important;
}

/* ══════════════════════════════════════════════
   BUTTONS
   ══════════════════════════════════════════════ */
.stButton > button {
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    border-radius: 10px !important;
    padding: 0.65rem 2rem !important;
    letter-spacing: 0.3px !important;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    border: none !important;
    position: relative;
    overflow: hidden;
}

.stButton > button[kind="primary"],
.stButton > button[data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg, #38BDF8, #0EA5E9) !important;
    color: #0B0F19 !important;
    font-weight: 700 !important;
    box-shadow: 0 4px 20px rgba(56, 189, 248, 0.25) !important;
}

.stButton > button[kind="primary"]:hover,
.stButton > button[data-testid="stBaseButton-primary"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 30px rgba(56, 189, 248, 0.35) !important;
}

.stButton > button[kind="secondary"],
.stButton > button[data-testid="stBaseButton-secondary"] {
    background: var(--bg-card) !important;
    color: var(--text-primary) !important;
    border: 1px solid var(--border-subtle) !important;
}

.stDownloadButton > button {
    background: linear-gradient(135deg, #34D399, #10B981) !important;
    color: #0B0F19 !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    padding: 0.65rem 2rem !important;
    border: none !important;
    box-shadow: 0 4px 20px rgba(52, 211, 153, 0.25) !important;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
}

.stDownloadButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 30px rgba(52, 211, 153, 0.35) !important;
}

/* ══════════════════════════════════════════════
   FILE UPLOADER
   ══════════════════════════════════════════════ */
[data-testid="stFileUploader"] {
    background: var(--bg-card) !important;
    border: 1px dashed var(--border-accent) !important;
    border-radius: 14px !important;
    padding: 1.5rem !important;
}

[data-testid="stFileUploader"] > label > div > p {
    color: var(--text-secondary) !important;
    font-size: 0.95rem !important;
    font-family: 'DM Sans', sans-serif !important;
}

[data-testid="stFileUploader"] section {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    gap: 0.5rem;
}

[data-testid="stFileUploader"] section > span {
    color: var(--text-muted) !important;
    font-size: 0.82rem !important;
}

[data-testid="stFileUploader"] button {
    min-height: 42px;
    padding: 0.55rem 1.5rem !important;
    background: var(--bg-card-hover) !important;
    color: var(--text-primary) !important;
    border: 1px solid var(--border-accent) !important;
    border-radius: 8px !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    position: relative !important;
    z-index: 2 !important;
    cursor: pointer !important;
}

[data-testid="stFileUploader"] button:hover {
    background: rgba(56, 189, 248, 0.15) !important;
    border-color: var(--accent-cyan) !important;
}

/* ══════════════════════════════════════════════
   DATA TABLES & DATAFRAMES — DARK THEME FIX
   ══════════════════════════════════════════════ */
[data-testid="stDataFrame"],
.stDataFrame {
    animation: scaleIn 0.5s ease-out;
}

[data-testid="stDataFrame"] [data-testid="glideDataEditor"],
.dvn-scroller {
    background: var(--bg-card) !important;
    border: 1px solid var(--border-subtle) !important;
    border-radius: 10px !important;
}

/* Theme the Glide Data Grid via its official CSS variables only.
   Do NOT use wildcard color overrides or force canvas opacity — Glide
   renders text onto <canvas> and those overrides cause blank cells. */
[data-testid="stDataFrame"] [data-testid="glideDataEditor"] {
    --gdg-bg-cell: #1A1F2E;
    --gdg-bg-header: #151A28;
    --gdg-text-dark: #F1F5F9;
    --gdg-text-medium: #94A3B8;
    --gdg-text-light: #64748B;
    --gdg-border-color: rgba(255,255,255,0.06);
    --gdg-bg-header-has-focus: #212840;
    --gdg-bg-cell-medium: #212840;
    --gdg-accent-color: #38BDF8;
    --gdg-accent-light: rgba(56, 189, 248, 0.15);
}

/* ══════════════════════════════════════════════
   ALERTS — Success, Info, Warning, Error
   ══════════════════════════════════════════════ */
[data-testid="stAlert"] {
    border-radius: 10px !important;
    border: none !important;
    animation: slideRight 0.5s ease-out;
}

.stAlert [data-testid="stAlertContentSuccess"],
div[data-baseweb="notification"][kind="positive"],
div.stSuccess {
    background: rgba(52, 211, 153, 0.08) !important;
    border-left: 3px solid var(--accent-emerald) !important;
}

.stAlert [data-testid="stAlertContentInfo"],
div[data-baseweb="notification"][kind="info"],
div.stInfo {
    background: rgba(56, 189, 248, 0.08) !important;
    border-left: 3px solid var(--accent-cyan) !important;
}

div.stWarning {
    background: rgba(251, 191, 36, 0.08) !important;
    border-left: 3px solid var(--accent-amber) !important;
}

div.stError {
    background: rgba(251, 113, 133, 0.08) !important;
    border-left: 3px solid var(--accent-rose) !important;
}

/* ══════════════════════════════════════════════
   EXPANDERS
   ══════════════════════════════════════════════ */
[data-testid="stExpander"] {
    background: var(--bg-card) !important;
    border: 1px solid var(--border-subtle) !important;
    border-radius: 10px !important;
    transition: all 0.3s ease !important;
}

[data-testid="stExpander"]:hover {
    border-color: var(--border-accent) !important;
}

[data-testid="stExpander"] summary > span:last-child {
    color: var(--text-primary) !important;
    font-weight: 500 !important;
    font-family: 'DM Sans', sans-serif !important;
}

[data-testid="stExpander"] summary > span:first-child:not(:last-child) {
    font-size: 0 !important;
    width: 1.2rem !important;
    height: 1.2rem !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
}

[data-testid="stExpander"] summary > span:first-child:not(:last-child)::before {
    content: '▸' !important;
    font-size: 1rem !important;
    font-family: sans-serif !important;
    color: var(--text-muted) !important;
}

[data-testid="stExpander"][open] summary > span:first-child:not(:last-child)::before {
    content: '▾' !important;
}

[data-testid="stExpander"] summary svg {
    flex-shrink: 0 !important;
}

/* ══════════════════════════════════════════════
   MULTISELECT / SELECT / INPUTS
   ══════════════════════════════════════════════ */
[data-testid="stMultiSelect"],
[data-testid="stSelectbox"] {
    animation: fadeIn 0.4s ease-out;
}

div[data-baseweb="select"] > div,
div[data-baseweb="input"] > div {
    background: var(--bg-input) !important;
    border-color: var(--border-subtle) !important;
    border-radius: 8px !important;
    color: var(--text-primary) !important;
}

div[data-baseweb="select"] > div:hover,
div[data-baseweb="input"] > div:hover {
    border-color: var(--accent-cyan) !important;
}

div[data-baseweb="tag"] {
    background: rgba(56, 189, 248, 0.15) !important;
    color: var(--accent-cyan) !important;
    border-radius: 6px !important;
}

div[data-baseweb="popover"] {
    background: var(--bg-card) !important;
    border: 1px solid var(--border-subtle) !important;
    border-radius: 10px !important;
}

div[data-baseweb="popover"] li {
    color: var(--text-secondary) !important;
}

div[data-baseweb="popover"] li:hover {
    background: var(--bg-card-hover) !important;
    color: var(--text-primary) !important;
}

/* ══════════════════════════════════════════════
   DIVIDERS
   ══════════════════════════════════════════════ */
hr, .stDivider, [data-testid="stDivider"] {
    border-color: var(--border-subtle) !important;
    opacity: 0.5;
}

/* ══════════════════════════════════════════════
   FEATURE CARDS
   ══════════════════════════════════════════════ */
.feature-card {
    background: var(--bg-card);
    border: 1px solid var(--border-subtle);
    border-radius: 14px;
    padding: 1.6rem;
    transition: all 0.35s cubic-bezier(0.4, 0, 0.2, 1);
    position: relative;
    overflow: hidden;
}

.feature-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, transparent, var(--accent-cyan), transparent);
    opacity: 0;
    transition: opacity 0.35s ease;
}

.feature-card:hover {
    border-color: var(--border-accent);
    transform: translateY(-4px);
    box-shadow: 0 12px 40px rgba(0,0,0,0.3);
}

.feature-card:hover::before {
    opacity: 1;
}

.feature-card .card-icon {
    font-size: 1.8rem;
    margin-bottom: 0.8rem;
    display: block;
}

.feature-card .card-title {
    color: var(--text-primary) !important;
    font-weight: 600;
    font-size: 1rem;
    margin-bottom: 0.5rem;
}

.feature-card .card-desc {
    color: var(--text-secondary) !important;
    font-size: 0.88rem;
    line-height: 1.6;
}

/* ══════════════════════════════════════════════
   DETECTION TABLE (on landing)
   ══════════════════════════════════════════════ */
.detect-table {
    width: 100%;
    border-collapse: separate;
    border-spacing: 0;
    background: var(--bg-card);
    border-radius: 12px;
    overflow: hidden;
    border: 1px solid var(--border-subtle);
    animation: scaleIn 0.6s ease-out 0.5s both;
}

.detect-table thead th {
    background: rgba(56, 189, 248, 0.08);
    color: var(--accent-cyan) !important;
    font-weight: 600;
    font-size: 0.8rem;
    text-transform: uppercase;
    letter-spacing: 1px;
    padding: 0.9rem 1.2rem;
    text-align: left;
    border-bottom: 1px solid var(--border-subtle);
}

.detect-table tbody tr {
    transition: background 0.2s ease;
}

.detect-table tbody tr:hover {
    background: var(--bg-card-hover);
}

.detect-table tbody td {
    padding: 0.75rem 1.2rem;
    color: var(--text-secondary) !important;
    font-size: 0.9rem;
    border-bottom: 1px solid var(--border-subtle);
    font-family: 'DM Sans', sans-serif;
}

.detect-table tbody tr:last-child td {
    border-bottom: none;
}

.detect-table .type-cell {
    color: var(--text-primary) !important;
    font-weight: 500;
}

.detect-table .original-cell {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.82rem;
    color: var(--accent-rose) !important;
}

.detect-table .masked-cell {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.82rem;
    color: var(--accent-emerald) !important;
}

/* ══════════════════════════════════════════════
   CUSTOM MARKDOWN TABLE OVERRIDE
   ══════════════════════════════════════════════ */
[data-testid="stMarkdownContainer"] table {
    background: var(--bg-card) !important;
    border-radius: 10px;
    overflow: hidden;
    width: 100%;
}

[data-testid="stTable"] {
    overflow-x: auto !important;
    max-width: 100% !important;
}

[data-testid="stTable"] table {
    font-size: 0.82rem !important;
}

[data-testid="stMarkdownContainer"] table th {
    background: rgba(56, 189, 248, 0.08) !important;
    color: var(--accent-cyan) !important;
    font-weight: 600 !important;
    border-bottom: 1px solid var(--border-subtle) !important;
    padding: 0.8rem 1rem !important;
}

[data-testid="stMarkdownContainer"] table td {
    color: var(--text-secondary) !important;
    border-bottom: 1px solid var(--border-subtle) !important;
    padding: 0.7rem 1rem !important;
    background: transparent !important;
}

[data-testid="stMarkdownContainer"] table tr:hover td {
    background: var(--bg-card-hover) !important;
}

/* ══════════════════════════════════════════════
   FOOTER
   ══════════════════════════════════════════════ */
.pro-footer {
    text-align: center;
    padding: 2.5rem 0 1rem 0;
    color: var(--text-secondary);
    font-size: 0.82rem;
    border-top: 1px solid var(--border-subtle);
    margin-top: 3rem;
    letter-spacing: 0.3px;
}

.pro-footer .footer-brand {
    background: linear-gradient(135deg, #38BDF8, #A78BFA);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    font-weight: 700;
    font-size: 0.9rem;
}

/* ══════════════════════════════════════════════
   SPINNERS
   ══════════════════════════════════════════════ */
.stSpinner > div {
    border-top-color: var(--accent-cyan) !important;
}

/* ══════════════════════════════════════════════
   HIDE STREAMLIT CHROME
   ══════════════════════════════════════════════ */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}

/* ══════════════════════════════════════════════
   CAPTIONS & SMALL TEXT
   ══════════════════════════════════════════════ */
[data-testid="stCaptionContainer"] p,
.stCaption {
    color: var(--text-secondary) !important;
    font-size: 0.82rem !important;
}

/* ══════════════════════════════════════════════
   SCROLLBAR
   ══════════════════════════════════════════════ */
::-webkit-scrollbar {
    width: 6px;
    height: 6px;
}

::-webkit-scrollbar-track {
    background: var(--bg-primary);
}

::-webkit-scrollbar-thumb {
    background: var(--text-muted);
    border-radius: 3px;
}

::-webkit-scrollbar-thumb:hover {
    background: var(--text-secondary);
}

/* ══════════════════════════════════════════════
   COLUMN GAP FIX
   ══════════════════════════════════════════════ */
[data-testid="stHorizontalBlock"] {
    gap: 1.2rem;
}

/* ══════════════════════════════════════════════
   STAGGERED FEATURE CARD ANIMATIONS
   ══════════════════════════════════════════════ */
[data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-child(1) .feature-card {
    animation: fadeSlideIn 0.5s ease-out 0.2s both;
}
[data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-child(2) .feature-card {
    animation: fadeSlideIn 0.5s ease-out 0.35s both;
}
[data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:nth-child(3) .feature-card {
    animation: fadeSlideIn 0.5s ease-out 0.5s both;
}

/* ══════════════════════════════════════════════
   METRIC / STAT CARDS
   ══════════════════════════════════════════════ */
.stat-pill {
    display: inline-flex;
    align-items: center;
    gap: 0.5rem;
    background: var(--bg-card);
    border: 1px solid var(--border-subtle);
    border-radius: 50px;
    padding: 0.5rem 1.2rem;
    font-size: 0.9rem;
    color: var(--text-secondary);
    font-weight: 500;
    transition: all 0.3s ease;
}

.stat-pill .num {
    color: var(--accent-cyan);
    font-weight: 700;
    font-family: 'JetBrains Mono', monospace;
}
</style>

<!-- Floating Particles (CSS-only) -->
<div class="particles-container">
    <div class="particle"></div>
    <div class="particle"></div>
    <div class="particle"></div>
    <div class="particle"></div>
    <div class="particle"></div>
    <div class="particle"></div>
    <div class="particle"></div>
    <div class="particle"></div>
    <div class="particle"></div>
    <div class="particle"></div>
</div>
""", unsafe_allow_html=True)


fake_en = Faker('en_GB')
fake_sv = Faker('sv_SE')


# ─────────────────────────────────────────────
#  NLP — lazy-loaded spaCy models
# ─────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def load_nlp_models():
    """Load spaCy models for English & Swedish name detection. Cached across reruns."""
    try:
        import spacy
        try:
            nlp_en = spacy.load("en_core_web_sm")
        except OSError:
            import subprocess
            subprocess.run(["python", "-m", "spacy", "download", "en_core_web_sm"], check=True)
            nlp_en = spacy.load("en_core_web_sm")

        try:
            nlp_sv = spacy.load("sv_core_news_sm")
        except OSError:
            import subprocess
            subprocess.run(["python", "-m", "spacy", "download", "sv_core_news_sm"], check=True)
            nlp_sv = spacy.load("sv_core_news_sm")

        return nlp_en, nlp_sv
    except Exception as e:
        st.warning(f"Error loading NLP models: {e}")
        return None, None


def mask_names_ner(text, mask_mode, token_counters, lang, nlp_en, nlp_sv, vault):
    """NLP-based name detection with vault support for referential integrity."""
    if not nlp_en or not nlp_sv:
        return text
    if len(text.strip()) < 2:
        return text

    persons_found = set()
    for nlp in [nlp_en, nlp_sv]:
        doc = nlp(text)
        for ent in doc.ents:
            if ent.label_ in ("PER", "PERSON"):
                persons_found.add((ent.start_char, ent.end_char, ent.text))

    if not persons_found:
        return text

    persons = sorted(persons_found, key=lambda x: x[0], reverse=True)
    result = text
    for start, end, name in persons:
        key = ("PERSON", name)
        if key in vault:
            replacement = vault[key]
        else:
            if mask_mode in ("Redacted (●●●●●)", "Maskerad (●●●●●)"):
                replacement = '●' * len(name)
            elif mask_mode in ("Fake Realistic Data", "Falsk realistisk data"):
                fake = fake_sv if any(c in name.lower() for c in ['å', 'ä', 'ö']) else fake_en
                replacement = fake.name()
            else:
                token_counters["PERSON"] = token_counters.get("PERSON", 0) + 1
                replacement = f"PERSON_{token_counters['PERSON']:03d}"
            vault[key] = replacement
        result = result[:start] + replacement + result[end:]

    return result


# ─────────────────────────────────────────────
#  UI TRANSLATIONS
# ─────────────────────────────────────────────
T = {
    "en": {
        "title": "🛡️ DataMask",
        "subtitle": "Upload an Excel or CSV file. We'll scan it, mask the sensitive data, and give you a clean file to download.",
        "settings": "Settings",
        "what_to_mask": "Detection Rules",
        "masking_style": "Output Style",
        "mask_mode_label": "How should we replace sensitive data?",
        "mask_modes": ["Redacted (●●●●●)", "Fake Realistic Data", "Token (e.g. EMAIL_001)"],
        "tip": "💡 **Tip:** 'Fake Realistic Data' keeps your file usable for testing while removing all real personal info.\n\n✨ Person names are auto-detected using AI (spaCy NER) in both English and Swedish.\n\n🔗 Same value always gets the same mask — joins and lookups still work.",
        "upload_label": "Upload your file (.xlsx or .csv)",
        # FIX: more honest privacy statement
        "upload_help": "Files are processed in memory only — nothing is written to disk or logged.",
        "file_loaded": "✅ File loaded: **{name}** — {rows} rows, {cols} columns",
        "large_file_warning": "⚠️ Large file ({rows:,} rows). Masking may take several minutes, especially with Person Names (NLP) enabled.",
        "choose_cols": "Choose columns to scan",
        "cols_caption": "All columns are scanned by default. Untick any you want to skip.",
        "cols_label": "Columns to scan:",
        "preview_label": "Preview original data (first 10 rows)",
        "run_button": "Run Masking",
        "warn_no_patterns": "Please select at least one pattern to mask.",
        "spinner": "Scanning and masking your data…",
        "progress_col": "Scanning column: **{col}** ({i}/{n})",
        "done": "✅ Done — **{n}** values masked across your file.",
        "no_changes": "ℹ️ No sensitive data was detected in the selected columns. Your file appears to be clean already.",
        "original_sample": "🔴 Original",
        "masked_sample": "🟢 Masked",
        "report_expander": "Masking Report — {n} changes",
        "download_header": "Download",
        "download_button": "⬇️ Download Masked Excel (3 tabs)",
        "download_csv_button": "⬇️ Download Masked CSV only",
        "download_report_button": "⬇️ Download Report CSV",
        "download_caption": "Contains 3 tabs: Original Data, Masked Data, and Masking Report.",
        "group_names": "Names",
        "group_contact": "Contact",
        "group_financial": "Financial",
        "group_identifiers": "Identifiers",
        "group_location": "Location",
        "group_other": "Other",
        "select_all": "✓ Select all",
        "select_none": "✗ Select none",
        "landing_info": "Upload a file to get started →",
        "landing_table_header": "### What does DataMask detect?",
        "landing_table_rows": [
            ("Person Name (NLP) ✨", "Erik Johansson", "James Smith",        "both"),
            ("Email",                "john@company.com","fake@email.com",    "both"),
            ("UK Phone",             "07891 234567",    "07234 567890",      "en"),
            ("UK Postcode",          "SW1A 1AA",        "M4 3AB",            "en"),
            ("NI Number",            "AB123456C",       "XY654321D",         "en"),
            ("Salary (£€$)",         "£45,000",         "£67,000",           "en"),
            ("Personnummer",         "19850312-1234",   "19920814-5678",     "sv"),
            ("Swedish Phone",        "070-123 45 67",   "073-456 78 90",     "sv"),
            ("Swedish Postcode",     "113 45",          "211 56",            "sv"),
            ("SEK Amount",           "45 000 kr",       "67 000 kr",         "sv"),
            ("Credit Card",          "4111 1111 1111 1111","5412 7534 2341 9876","both"),
            ("Date of Birth",        "12/05/1987",      "23/08/1994",        "both"),
            ("IP Address",           "192.168.1.1",     "83.21.45.7",        "both"),
        ],
        "file_error": "Couldn't read that file. Error: {e}",
        "csv_encoding_error": "Couldn't decode this CSV file. Try saving it as UTF-8 or Excel (.xlsx).",
        "trust_secure": "In-memory only",
        "trust_client": "Session-isolated",
        "trust_gdpr": "GDPR compliant",
        "trust_zero": "Zero data stored",
        "why_title": "Why DataMask?",
        "card_1_title": "In-Memory Processing",
        "card_1_desc": "Your data is processed in memory only. Nothing is persisted to disk or logged.",
        "card_2_title": "AI-Powered Detection",
        "card_2_desc": "Advanced NLP recognizes names in English & Swedish automatically.",
        "card_3_title": "Referential Integrity",
        "card_3_desc": "Same value always maps to the same mask — joins and lookups still work.",
        "pattern_col": "Pattern",
    },
    "sv": {
        "title": "🛡️ DataMask",
        "subtitle": "Ladda upp en Excel- eller CSV-fil. Vi skannar, maskerar känsliga uppgifter och ger dig en ren fil.",
        "settings": "Inställningar",
        "what_to_mask": "Detektionsregler",
        "masking_style": "Utdatastil",
        "mask_mode_label": "Hur ska känsliga uppgifter ersättas?",
        "mask_modes": ["Maskerad (●●●●●)", "Falsk realistisk data", "Token (t.ex. EMAIL_001)"],
        "tip": "💡 **Tips:** 'Falsk realistisk data' håller filen användbar för testning.\n\n✨ Personnamn identifieras automatiskt med AI (spaCy NER) på engelska och svenska.\n\n🔗 Samma värde får alltid samma mask — join och lookups fungerar fortfarande.",
        "upload_label": "Ladda upp din fil (.xlsx eller .csv)",
        "upload_help": "Filer bearbetas endast i minnet — inget skrivs till disk eller loggas.",
        "file_loaded": "✅ Inläst: **{name}** — {rows} rader, {cols} kolumner",
        "large_file_warning": "⚠️ Stor fil ({rows:,} rader). Maskering kan ta flera minuter, särskilt med Personnamn (NLP) aktiverat.",
        "choose_cols": "Välj kolumner att skanna",
        "cols_caption": "Alla kolumner skannas som standard. Avmarkera de du vill hoppa över.",
        "cols_label": "Kolumner att skanna:",
        "preview_label": "Förhandsgranska (första 10 raderna)",
        "run_button": "Kör maskering",
        "warn_no_patterns": "Välj minst ett mönster.",
        "spinner": "Skannar och maskerar…",
        "progress_col": "Skannar kolumn: **{col}** ({i}/{n})",
        "done": "✅ Klart — **{n}** värden maskerades.",
        "no_changes": "ℹ️ Ingen känslig data hittades i de valda kolumnerna. Filen verkar redan vara ren.",
        "original_sample": "🔴 Original",
        "masked_sample": "🟢 Maskerad",
        "report_expander": "Maskeringsrapport — {n} ändringar",
        "download_header": "Ladda ner",
        "download_button": "⬇️ Ladda ner maskerad Excel (3 flikar)",
        "download_csv_button": "⬇️ Ladda ner bara maskerad CSV",
        "download_report_button": "⬇️ Ladda ner rapport som CSV",
        "download_caption": "Filen har 3 flikar: Originaldata, Maskerad data och Rapport.",
        "group_names": "Namn",
        "group_contact": "Kontakt",
        "group_financial": "Finansiellt",
        "group_identifiers": "ID-nummer",
        "group_location": "Plats",
        "group_other": "Övrigt",
        "select_all": "✓ Välj alla",
        "select_none": "✗ Välj inga",
        "landing_info": "Ladda upp en fil för att börja →",
        "landing_table_header": "### Vad identifierar DataMask?",
        "landing_table_rows": [
            ("Personnamn (NLP) ✨",  "Erik Johansson",          "Lars Svensson",             "both"),
            ("E-post",               "kalle@foretag.se",        "falsk@epost.se",             "both"),
            ("Personnummer",         "19850312-1234",           "19920814-5678",              "sv"),
            ("Svenskt mobilnummer",  "070-123 45 67",           "073-456 78 90",              "sv"),
            ("Svenskt postnummer",   "113 45",                  "211 56",                     "sv"),
            ("SEK-belopp",           "45 000 kr",               "67 000 kr",                  "sv"),
            ("Brittiskt telefon",    "07891 234567",            "07234 567890",               "en"),
            ("Brittiskt postnummer", "SW1A 1AA",                "M4 3AB",                     "en"),
            ("NI-nummer (UK)",       "AB123456C",               "XY654321D",                  "en"),
            ("Lön (£€$)",            "£45,000",                 "£67,000",                    "en"),
            ("Kreditkort",           "4111 1111 1111 1111",     "5412 7534 2341 9876",        "both"),
            ("Födelsedatum",         "12/05/1987",              "23/08/1994",                 "both"),
            ("IP-adress",            "192.168.1.1",             "83.21.45.7",                 "both"),
        ],
        "file_error": "Kunde inte läsa filen. Fel: {e}",
        "csv_encoding_error": "Kunde inte avkoda CSV-filen. Spara om den som UTF-8 eller Excel (.xlsx).",
        "trust_secure": "Endast i minnet",
        "trust_client": "Sessionisolerad",
        "trust_gdpr": "GDPR-kompatibel",
        "trust_zero": "Ingen lagring",
        "why_title": "Varför DataMask?",
        "card_1_title": "Bearbetning i minnet",
        "card_1_desc": "Din data bearbetas endast i minnet. Inget sparas till disk eller loggas.",
        "card_2_title": "AI-driven identifiering",
        "card_2_desc": "Avancerad NLP känner igen namn på engelska och svenska automatiskt.",
        "card_3_title": "Referentiell integritet",
        "card_3_desc": "Samma värde får alltid samma mask — join och lookups fungerar fortfarande.",
        "pattern_col": "Mönster",
    },
}


# ─────────────────────────────────────────────
#  REGEX PATTERNS
#  Order matters: most-specific → most-general.
# ─────────────────────────────────────────────
PATTERNS = {
    "Person Names (NLP)":     None,

    # FIX #1 — Email regex now supports Unicode word characters (å ä ö etc.)
    # via \w which in Python 3 re matches Unicode letters by default.
    # The old pattern [a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+ silently
    # skipped Swedish emails like 'åsa@företag.se'.
    "Email Addresses":        r'[\w.+\-]+@[\w\-]+\.[\w.\-]+',

    "Credit Card Numbers":    r'\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4}',
    "Swedish Personnummer":   r'\b(?:19|20)?\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])[-\s]?\d{4}\b',
    "National Insurance":     r'\b[A-Z]{2}\s?\d{2}\s?\d{2}\s?\d{2}\s?[A-Z]\b',

    # FIX #2 — IP regex now validates each octet is 0-255, so random version
    # strings like 'v1.2.3.4' and '999.999.999.999' are no longer matched.
    "IP Addresses":           r'\b(?:25[0-5]|2[0-4]\d|1\d\d|[1-9]?\d)(?:\.(?:25[0-5]|2[0-4]\d|1\d\d|[1-9]?\d)){3}\b',

    "UK Postcodes":           r'(?i:\b[A-Z]{1,2}\d[A-Z\d]?\s?\d[A-Z]{2}\b)',
    "UK Phone Numbers":       r'(?:\+44\s?|0)7\d{3}[\s-]?\d{3}[\s-]?\d{3}|(?:\+44\s?|0)7\d{3}[\s-]?\d{6}',
    "Swedish Phone Numbers":  r'(?:\+46[\s-]?|0)7[0-9][\s-]?\d{3}[\s-]?\d{2}[\s-]?\d{2}',
    "US Phone Numbers":       r'\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}',
    "Dates of Birth":         r'\b\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}\b',
    "Salary / Currency":      r'[£€$]\s?\d{1,3}(?:[,\s]\d{3})*(?:\.\d{2})?',
    "SEK Currency":           r'\d{1,3}(?:\s\d{3})+\s*kr',

    # FIX #3 — Swedish postcodes: first digit 1-9 (no leading 0)
    # Reduces false positives on SKUs/IDs a bit (can't eliminate entirely
    # without context — '113 45' is genuinely ambiguous in isolation).
    "Swedish Postcodes":      r'\b[1-9]\d{2}\s\d{2}\b',
}

PATTERN_LANG = {k: "both" for k in PATTERNS}

PATTERN_LABELS = {
    "en": {
        "Person Names (NLP)":    "Person Names (NLP) ✨",
        "Email Addresses":       "Email Addresses",
        "UK Phone Numbers":      "UK Phone Numbers",
        "US Phone Numbers":      "US Phone Numbers",
        "UK Postcodes":          "UK Postcodes",
        "National Insurance":    "National Insurance (NI)",
        "Swedish Personnummer":  "Swedish Personnummer",
        "Swedish Phone Numbers": "Swedish Phone Numbers",
        "Swedish Postcodes":     "Swedish Postcodes",
        "Credit Card Numbers":   "Credit Card Numbers",
        "Dates of Birth":        "Dates of Birth",
        "IP Addresses":          "IP Addresses",
        "Salary / Currency":     "Salary / Currency (£€$)",
        "SEK Currency":          "SEK Currency (kr)",
    },
    "sv": {
        "Person Names (NLP)":    "Personnamn (NLP) ✨",
        "Email Addresses":       "E-postadresser",
        "UK Phone Numbers":      "Brittiska telefonnummer",
        "US Phone Numbers":      "Amerikanska telefonnummer",
        "UK Postcodes":          "Brittiska postnummer",
        "National Insurance":    "NI-nummer (UK)",
        "Swedish Personnummer":  "Personnummer",
        "Swedish Phone Numbers": "Svenska mobilnummer",
        "Swedish Postcodes":     "Svenska postnummer",
        "Credit Card Numbers":   "Kreditkortsnummer",
        "Dates of Birth":        "Födelsedatum",
        "IP Addresses":          "IP-adresser",
        "Salary / Currency":     "Lön / Valuta (£€$)",
        "SEK Currency":          "SEK-belopp (kr)",
    },
}

# Pattern groupings for cleaner sidebar layout
PATTERN_GROUPS = {
    "group_names":       ["Person Names (NLP)"],
    "group_contact":     ["Email Addresses", "UK Phone Numbers", "US Phone Numbers", "Swedish Phone Numbers"],
    "group_financial":   ["Salary / Currency", "SEK Currency", "Credit Card Numbers"],
    "group_identifiers": ["Swedish Personnummer", "National Insurance", "Dates of Birth"],
    "group_location":    ["UK Postcodes", "Swedish Postcodes", "IP Addresses"],
}


# ─────────────────────────────────────────────
#  FAKE DATA REPLACEMENTS
# ─────────────────────────────────────────────
def replace_with_fake(category, lang):
    fake = fake_sv if lang == "sv" else fake_en

    def fake_personnummer():
        dob = fake_sv.date_of_birth(minimum_age=18, maximum_age=80)
        last4 = f"{random.randint(1000, 9999)}"
        return dob.strftime('%Y%m%d') + '-' + last4

    def fake_swedish_phone():
        prefixes = ['070', '072', '073', '076', '079']
        prefix = random.choice(prefixes)
        number = f"{random.randint(100, 999)} {random.randint(10, 99)} {random.randint(10, 99)}"
        return f"{prefix}-{number}"

    def fake_swedish_postcode():
        return f"{random.randint(100, 999)} {random.randint(10, 99):02d}"

    def fake_sek():
        amount = random.randint(10000, 120000)
        formatted = f"{amount:,}".replace(",", " ")
        return f"{formatted} kr"

    replacements = {
        "Email Addresses":       fake.email,
        "UK Phone Numbers":      fake_en.phone_number,
        "US Phone Numbers":      fake_en.phone_number,
        "UK Postcodes":          fake_en.postcode,
        "National Insurance":    lambda: f"{fake_en.lexify('??').upper()}{fake_en.numerify('######')}{fake_en.lexify('?').upper()}",
        "Swedish Personnummer":  fake_personnummer,
        "Swedish Phone Numbers": fake_swedish_phone,
        "Swedish Postcodes":     fake_swedish_postcode,
        "Credit Card Numbers":   lambda: fake.credit_card_number(card_type=None),
        "Dates of Birth":        lambda: fake.date_of_birth(minimum_age=18, maximum_age=80).strftime('%d/%m/%Y'),
        "IP Addresses":          fake.ipv4,
        "Salary / Currency":     lambda: f"£{fake_en.random_int(20000, 120000):,}",
        "SEK Currency":          fake_sek,
    }
    fn = replacements.get(category)
    return fn() if fn else "****"


# ─────────────────────────────────────────────
#  CORE MASKING LOGIC
# ─────────────────────────────────────────────
def _clean_token_key(category):
    """Convert category name to a clean TOKEN key."""
    key = re.sub(r'[^A-Za-z0-9]+', '_', category).strip('_').upper()
    return key


def mask_cell(value, selected_patterns, mask_mode, token_counters, lang, vault,
              nlp_en=None, nlp_sv=None, category_hits=None):
    """
    Mask sensitive data in a single cell value.

    vault : dict — maps (category, original_string) → replacement_string.
                   Ensures the same input always produces the same output.
    category_hits : optional dict — if provided, records which categories matched.
    """
    text = str(value)
    use_nlp_for_names = "Person Names (NLP)" in selected_patterns

    for category, pattern in PATTERNS.items():
        if category not in selected_patterns:
            continue
        if category == "Person Names (NLP)":
            continue

        def replace_match(m, c=category):
            original = m.group()
            key = (c, original)
            if category_hits is not None:
                category_hits[original] = c
            if key in vault:
                return vault[key]

            if mask_mode in ("Redacted (●●●●●)", "Maskerad (●●●●●)"):
                replacement = '●' * len(original)
            elif mask_mode in ("Fake Realistic Data", "Falsk realistisk data"):
                replacement = replace_with_fake(c, lang)
            else:  # token mode
                key_name = _clean_token_key(c)
                token_counters[key_name] = token_counters.get(key_name, 0) + 1
                replacement = f"{key_name}_{token_counters[key_name]:03d}"

            vault[key] = replacement
            return replacement

        text = re.sub(pattern, replace_match, text)

    if use_nlp_for_names and nlp_en and nlp_sv:
        before = text
        text = mask_names_ner(text, mask_mode, token_counters, lang, nlp_en, nlp_sv, vault)
        if category_hits is not None and text != before:
            category_hits["__nlp__"] = "Person Names (NLP)"

    return text


def process_dataframe(df, selected_patterns, mask_mode, selected_columns, lang,
                      progress_callback=None):
    """Main masking loop. Calls progress_callback(i, n, col) per column."""
    masked_df = df.copy()
    token_counters = {}
    vault = {}
    report = []

    nlp_en, nlp_sv = (None, None)
    if "Person Names (NLP)" in selected_patterns:
        nlp_en, nlp_sv = load_nlp_models()

    cols_to_process = [c for c in (selected_columns or df.columns.tolist()) if c in df.columns]
    n = len(cols_to_process)

    for i, col in enumerate(cols_to_process):
        if progress_callback:
            progress_callback(i, n, col)

        for idx, value in df[col].items():
            if pd.isna(value):
                continue
            original = str(value)
            category_hits = {}
            cleaned = mask_cell(
                original, selected_patterns, mask_mode, token_counters, lang,
                vault=vault, nlp_en=nlp_en, nlp_sv=nlp_sv, category_hits=category_hits,
            )
            if cleaned != original:
                masked_df.at[idx, col] = cleaned
                # Best-effort: first category that matched (or __nlp__)
                pattern_label = next(iter(category_hits.values()), "—")
                report.append({
                    "Row": idx + 2,
                    "Column": col,
                    "Pattern": pattern_label,
                    "Original": original,
                    "Masked As": cleaned,
                })

    if progress_callback:
        progress_callback(n, n, "")

    return masked_df, pd.DataFrame(report)


# ─────────────────────────────────────────────
#  CSV LOADER (encoding fallback)
# ─────────────────────────────────────────────
def load_csv_with_fallback(uploaded_file):
    """Try common encodings in order."""
    encodings = ("utf-8", "utf-8-sig", "cp1252", "iso-8859-1", "latin-1")
    last_err = None
    for enc in encodings:
        try:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, encoding=enc)
        except (UnicodeDecodeError, UnicodeError) as e:
            last_err = e
            continue
    raise ValueError(f"Could not decode CSV with any common encoding. Last error: {last_err}")


# ─────────────────────────────────────────────
#  EXPORT HELPERS
# ─────────────────────────────────────────────
def export_to_excel(original_df, masked_df, report_df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        original_df.to_excel(writer, sheet_name='Original Data', index=False)
        masked_df.to_excel(writer, sheet_name='Masked Data', index=False)
        if not report_df.empty:
            report_df.to_excel(writer, sheet_name='Masking Report', index=False)

        wb = writer.book
        ws = wb['Masked Data']
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        for cell in ws[1]:
            cell.fill = green_fill
            cell.font = Font(bold=True)

    output.seek(0)
    return output


def export_to_csv(df):
    return df.to_csv(index=False).encode('utf-8-sig')


# ═════════════════════════════════════════════
#  UI — STREAMLIT INTERFACE
# ═════════════════════════════════════════════

# Session state init
for key in ('masked_df', 'report_df', 'original_df', 'last_uploaded_name'):
    if key not in st.session_state:
        st.session_state[key] = None

# ── Sidebar: language selector ──────────────
with st.sidebar:
    lang = st.radio(
        "Language / Språk",
        options=["en", "sv"],
        format_func=lambda x: "🇬🇧 English" if x == "en" else "🇸🇪 Svenska",
        horizontal=True,
        key="language_selector"
    )
    st.divider()

ui = T[lang]
labels = PATTERN_LABELS[lang]

# ── Sidebar: settings ───────────────────────
with st.sidebar:
    st.header(ui["settings"])
    st.subheader(ui["what_to_mask"])

    # Quick select/deselect buttons
    col_a, col_b = st.columns(2)
    with col_a:
        if st.button(ui["select_all"], use_container_width=True, key="btn_select_all"):
            for cat in PATTERNS:
                st.session_state[f"chk_{cat}"] = True
            st.rerun()
    with col_b:
        if st.button(ui["select_none"], use_container_width=True, key="btn_select_none"):
            for cat in PATTERNS:
                st.session_state[f"chk_{cat}"] = False
            st.rerun()

    selected_patterns = []
    # Grouped checkboxes
    for group_key, categories in PATTERN_GROUPS.items():
        st.markdown(f'<span class="group-label">{ui[group_key]}</span>', unsafe_allow_html=True)
        for category in categories:
            if category not in PATTERNS:
                continue
            default_val = st.session_state.get(f"chk_{category}", True)
            if st.checkbox(labels[category], value=default_val, key=f"chk_{category}"):
                selected_patterns.append(category)

    st.subheader(ui["masking_style"])
    mask_mode = st.radio(
        ui["mask_mode_label"],
        ui["mask_modes"],
        index=1,
        key="mask_mode_radio"
    )

    # FIX #4 — Compare by INDEX, not by localized string, so switching
    # languages doesn't wipe the user's results.
    current_mode_idx = ui["mask_modes"].index(mask_mode)
    if st.session_state.get('previous_mode_idx') is None:
        st.session_state.previous_mode_idx = current_mode_idx
    elif st.session_state.previous_mode_idx != current_mode_idx:
        st.session_state.masked_df = None
        st.session_state.report_df = None
        st.session_state.original_df = None
        st.session_state.previous_mode_idx = current_mode_idx


# ── Hero Header ──────────────────────────────
st.title(ui["title"])
st.markdown(f'<p class="hero-subtitle">{ui["subtitle"]}</p>', unsafe_allow_html=True)

# Trust banner
st.markdown(f'''
<div class="trust-banner">
    <span class="trust-badge"><span class="icon">🔐</span> {ui["trust_secure"]}</span>
    <span class="trust-badge"><span class="icon">💻</span> {ui["trust_client"]}</span>
    <span class="trust-badge"><span class="icon">✅</span> {ui["trust_gdpr"]}</span>
    <span class="trust-badge"><span class="icon">🚫</span> {ui["trust_zero"]}</span>
</div>
''', unsafe_allow_html=True)


# ── File Upload ──────────────────────────────
uploaded_file = st.file_uploader(
    ui["upload_label"],
    type=["xlsx", "csv"],
    help=ui["upload_help"],
    label_visibility="visible",
)

if uploaded_file:
    # Clear results if a different file is uploaded
    if st.session_state.last_uploaded_name != uploaded_file.name:
        st.session_state.masked_df = None
        st.session_state.report_df = None
        st.session_state.original_df = None
        st.session_state.last_uploaded_name = uploaded_file.name

    try:
        if uploaded_file.name.lower().endswith(".csv"):
            try:
                df = load_csv_with_fallback(uploaded_file)
            except ValueError:
                st.error(ui["csv_encoding_error"])
                st.stop()
        else:
            df = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(ui["file_error"].format(e=e))
        st.stop()

    st.success(ui["file_loaded"].format(name=uploaded_file.name, rows=len(df), cols=len(df.columns)))

    if len(df) > 20_000:
        st.warning(ui["large_file_warning"].format(rows=len(df)))

    st.divider()

    st.subheader(ui["choose_cols"])
    st.caption(ui["cols_caption"])
    all_cols = df.columns.tolist()
    selected_columns = st.multiselect(
        ui["cols_label"],
        options=all_cols,
        default=all_cols,
    )

    # FIX #5 — Use st.dataframe for previews (scrollable, sortable, resizable)
    with st.expander(ui["preview_label"], expanded=False):
        st.dataframe(df.head(10), use_container_width=True, hide_index=True)

    st.divider()

    if st.button(ui["run_button"], type="primary", use_container_width=True):
        if not selected_patterns:
            st.warning(ui["warn_no_patterns"])
        else:
            # Pre-load NLP models with a clear spinner
            if "Person Names (NLP)" in selected_patterns:
                with st.spinner("Loading AI models…" if lang == "en" else "Laddar AI-modeller…"):
                    nlp_en, nlp_sv = load_nlp_models()
                    if not nlp_en or not nlp_sv:
                        st.error("⚠️ NLP models not available. Person names won't be masked.")

            # FIX #6 — Real progress bar per column instead of indefinite spinner
            progress_bar = st.progress(0.0, text=ui["spinner"])

            def update_progress(i, n, col):
                frac = i / n if n else 1.0
                text = ui["progress_col"].format(col=col, i=i, n=n) if col else ui["spinner"]
                progress_bar.progress(frac, text=text)

            masked_df, report_df = process_dataframe(
                df, selected_patterns, mask_mode, selected_columns, lang,
                progress_callback=update_progress,
            )
            progress_bar.empty()

            st.session_state.masked_df = masked_df
            st.session_state.report_df = report_df
            st.session_state.original_df = df

    # ── Results ──────────────────────────────
    if st.session_state.masked_df is not None:
        n_changes = len(st.session_state.report_df)
        if n_changes == 0:
            st.info(ui["no_changes"])
        else:
            st.success(ui["done"].format(n=n_changes))
        st.divider()

        col1, col2 = st.columns(2)
        with col1:
            st.subheader(ui["original_sample"])
            st.dataframe(st.session_state.original_df.head(10),
                         use_container_width=True, hide_index=True)
        with col2:
            st.subheader(ui["masked_sample"])
            st.dataframe(st.session_state.masked_df.head(10),
                         use_container_width=True, hide_index=True)

        if not st.session_state.report_df.empty:
            with st.expander(ui["report_expander"].format(n=len(st.session_state.report_df))):
                st.dataframe(
                    st.session_state.report_df.head(200),
                    use_container_width=True, hide_index=True,
                )

        st.divider()
        st.subheader(ui["download_header"])

        original_name = uploaded_file.name.rsplit('.', 1)[0]
        excel_output = export_to_excel(
            st.session_state.original_df,
            st.session_state.masked_df,
            st.session_state.report_df,
        )

        # FIX #7 — Multiple download formats (Excel + CSV + report-only)
        dl_col1, dl_col2, dl_col3 = st.columns(3)
        with dl_col1:
            st.download_button(
                label=ui["download_button"],
                data=excel_output,
                file_name=f"{original_name}_MASKED.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        with dl_col2:
            st.download_button(
                label=ui["download_csv_button"],
                data=export_to_csv(st.session_state.masked_df),
                file_name=f"{original_name}_MASKED.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with dl_col3:
            if not st.session_state.report_df.empty:
                st.download_button(
                    label=ui["download_report_button"],
                    data=export_to_csv(st.session_state.report_df),
                    file_name=f"{original_name}_REPORT.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

        st.caption(ui["download_caption"])

else:
    # ── Landing Page ─────────────────────────
    st.info(ui["landing_info"])
    st.markdown(ui["landing_table_header"])

    label_to_category = {
        "Person Name (NLP) ✨": "PERSON",
        "Personnamn (NLP) ✨": "PERSON",
        "Email": "Email Addresses",
        "E-post": "Email Addresses",
        "UK Phone": "UK Phone Numbers",
        "Brittiskt telefon": "UK Phone Numbers",
        "UK Postcode": "UK Postcodes",
        "Brittiskt postnummer": "UK Postcodes",
        "NI Number": "National Insurance",
        "NI-nummer (UK)": "National Insurance",
        "Salary (£€$)": "Salary / Currency",
        "Lön (£€$)": "Salary / Currency",
        "Personnummer": "Swedish Personnummer",
        "Swedish Phone": "Swedish Phone Numbers",
        "Svenskt mobilnummer": "Swedish Phone Numbers",
        "Swedish Postcode": "Swedish Postcodes",
        "Svenskt postnummer": "Swedish Postcodes",
        "SEK Amount": "SEK Currency",
        "SEK-belopp": "SEK Currency",
        "Credit Card": "Credit Card Numbers",
        "Kreditkort": "Credit Card Numbers",
        "Date of Birth": "Dates of Birth",
        "Födelsedatum": "Dates of Birth",
        "IP Address": "IP Addresses",
        "IP-adress": "IP Addresses",
    }

    rows = ui["landing_table_rows"]

    th_type = "Type" if lang == "en" else "Typ"
    th_example = "Example" if lang == "en" else "Exempel"
    th_masked = "Masked As" if lang == "en" else "Maskeras som"

    table_html = f'''<table class="detect-table">
    <thead><tr>
        <th>{th_type}</th>
        <th>{th_example}</th>
        <th>{th_masked}</th>
    </tr></thead><tbody>'''

    token_counters_demo = {}
    vault_demo = {}
    for label, example, _, row_lang in rows:
        if row_lang in (lang, "both"):
            category = label_to_category.get(label)
            if category == "PERSON":
                if mask_mode in ("Redacted (●●●●●)", "Maskerad (●●●●●)"):
                    masked_example = '●' * len(example)
                elif mask_mode in ("Fake Realistic Data", "Falsk realistisk data"):
                    fake = fake_sv if lang == "sv" else fake_en
                    masked_example = fake.name()
                else:
                    token_counters_demo["PERSON"] = token_counters_demo.get("PERSON", 0) + 1
                    masked_example = f"PERSON_{token_counters_demo['PERSON']:03d}"
            elif category and category in PATTERNS:
                masked_example = mask_cell(
                    example, [category], mask_mode, token_counters_demo, lang,
                    vault=vault_demo,
                )
            else:
                masked_example = example

            table_html += f'''<tr>
                <td class="type-cell">{label}</td>
                <td class="original-cell">{example}</td>
                <td class="masked-cell">{masked_example}</td>
            </tr>'''

    table_html += '</tbody></table>'
    st.markdown(table_html, unsafe_allow_html=True)

    # ── Feature Cards ────────────────────────
    st.markdown("---")
    st.markdown(f"### {ui['why_title']}")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f'''<div class="feature-card">
            <span class="card-icon">🔒</span>
            <div class="card-title">{ui["card_1_title"]}</div>
            <div class="card-desc">{ui["card_1_desc"]}</div>
        </div>''', unsafe_allow_html=True)
    with col2:
        st.markdown(f'''<div class="feature-card">
            <span class="card-icon">🤖</span>
            <div class="card-title">{ui["card_2_title"]}</div>
            <div class="card-desc">{ui["card_2_desc"]}</div>
        </div>''', unsafe_allow_html=True)
    with col3:
        st.markdown(f'''<div class="feature-card">
            <span class="card-icon">🔗</span>
            <div class="card-title">{ui["card_3_title"]}</div>
            <div class="card-desc">{ui["card_3_desc"]}</div>
        </div>''', unsafe_allow_html=True)


# ── Footer ───────────────────────────────────
footer_text = "Privacy-first data anonymization" if lang == "en" else "Integritetsfokuserad dataanonymisering"
st.markdown(f'''
<div class="pro-footer">
    <span class="footer-brand">DataMask</span> · {footer_text} · © 2026
</div>
''', unsafe_allow_html=True)
